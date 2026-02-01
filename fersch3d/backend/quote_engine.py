"""
Quote engine for Fersch 3D on‑demand printing service.

This module reads the pricing parameters from the provided Excel file
(`Calcul prix 3D FERSCH.xlsx`) and exposes functions to compute a
detailed cost breakdown given the volume of a part and selected options
such as material, type of piece, typology (bag size), and shipping.

The calculation follows the formulas defined in the original Excel:

* Material cost = volume × price/ml (with loss)
* Machine cost = volume × machine_time_per_ml × machine_hour_rate
* Base cost = material + machine
* Post‑treatment cost = base × post_rate
* Finishing cost = base × finish_rate (basic sanding included)
* Total cost before mark‑up = base + post + finish
* Mark‑up factor depends on volume thresholds
* Price HT per plate = total cost × mark‑up factor
* Packaging cost depends on total volume thresholds
* Bag cost depends on typology (Standard, Fragile, Grosse pièce)
* Shipping cost: 0 for retrait, fixed amount for delivery
* Total HT = price_ht_plate + bag + packaging + shipping
* TVA = total_ht × tva_rate
* Total TTC = total_ht + tva

The module caches the parameters on first load for efficiency. If the
Excel file changes, reinstantiate the QuoteEngine.

Author: Fersch 3D quoting engine
"""

from __future__ import annotations

import threading
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl


@dataclass
class Material:
    name: str
    price_per_ml_with_loss: float


@dataclass
class TypePiece:
    name: str
    factor: float


@dataclass
class Typology:
    name: str
    bag_price: float


class QuoteEngine:
    """Loads pricing parameters from an Excel file and computes quotes."""

    def __init__(self, excel_path: str | Path):
        self.excel_path = Path(excel_path)
        self._lock = threading.Lock()
        self._loaded = False
        # Containers for parameters
        self.materials: Dict[str, Material] = {}
        self.type_pieces: Dict[str, TypePiece] = {}
        self.typologies: Dict[str, Typology] = {}
        self.markup_table: List[Tuple[float, float]] = []  # (ml_min, factor)
        self.packaging_table: List[Tuple[float, float]] = []  # (ml_total_min, price)
        self.post_rate: float = 0.0
        self.finish_rate: float = 0.0
        self.tva_rate: float = 0.0
        self.machine_hour_rate: float = 7.0  # default if not loaded
        self.machine_time_per_ml: float = 0.0687  # hours per ml, approximated from PreForm
        self.shipping_retrait: float = 0.0
        self.shipping_delivery: float = 12.0
        # Additional parameters for supports and print time
        # Percentage of supports to add to the base volume per material
        self.material_support_percent: Dict[str, float] = {}
        # Print speed in tenth of mm/s per material (e.g. 10 -> 1 mm/s)
        self.material_print_speed: Dict[str, float] = {}
        # Global time factor for print time estimation
        self.time_factor: float = 1.0
        # Populate parameters
        self._load_parameters()
        # After loading parameters, apply defaults from environment variables if provided
        import os
        try:
            default_support = float(os.getenv('DEFAULT_SUPPORT_PERCENT', '0'))
        except ValueError:
            default_support = 0.0
        try:
            default_speed = float(os.getenv('DEFAULT_PRINT_SPEED', '10'))  # tenth mm/s
        except ValueError:
            default_speed = 10.0
        try:
            self.time_factor = float(os.getenv('TIME_FACTOR', '1'))
        except ValueError:
            self.time_factor = 1.0
        # Populate per-material defaults if not specified
        for mat_name in self.materials:
            self.material_support_percent.setdefault(mat_name, default_support)
            self.material_print_speed.setdefault(mat_name, default_speed)

    def _load_parameters(self) -> None:
        """Internal method to read the Excel file and extract parameters."""
        with self._lock:
            if self._loaded:
                return
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            ws = wb['Paramètres']
            # Load materials
            # Materials are listed starting at row 7 (1‑indexed) under the header row (row 6)
            row = 7
            while True:
                name_cell = ws.cell(row=row, column=1).value
                if not name_cell or not isinstance(name_cell, str):
                    break
                name = name_cell.strip()
                price_per_litre = ws.cell(row=row, column=2).value or 0.0
                loss_percent = ws.cell(row=row, column=3).value or 0.0
                # Compute price per ml with loss; handle header strings gracefully
                raw_price = ws.cell(row=row, column=5).value
                price_per_ml_with_loss: float
                try:
                    price_per_ml_with_loss = float(raw_price)
                except (TypeError, ValueError):
                    # If the cell contains text (header), compute manually
                    price_per_ml = (price_per_litre / 1000.0) if price_per_litre else 0.0
                    price_per_ml_with_loss = price_per_ml * (1.0 + float(loss_percent))
                self.materials[name] = Material(name=name, price_per_ml_with_loss=price_per_ml_with_loss)
                row += 1
            # Load machine cost
            # 'Coût horaire arrondi (€/h)' is at row 10 column 8
            try:
                machine_hour = ws.cell(row=10, column=8).value
                if machine_hour:
                    self.machine_hour_rate = float(machine_hour)
            except Exception:
                pass
            # Load type pieces, typologies, and rates
            # Skip header rows (row 12 and 13) then read until we hit a blank row or the shipping section
            row = 14  # first row containing actual type definitions
            while row <= ws.max_row:
                type_name = ws.cell(row=row, column=1).value
                # Stop if we reach a row where column 1 is None and column 7 has 'RETRAIT / LIVRAISON'
                if type_name is None and ws.cell(row=row, column=7).value:
                    break
                if type_name and isinstance(type_name, str):
                    try:
                        factor = float(ws.cell(row=row, column=2).value)
                        self.type_pieces[type_name.strip()] = TypePiece(name=type_name.strip(), factor=factor)
                    except (TypeError, ValueError):
                        pass
                typology_name = ws.cell(row=row, column=4).value
                if typology_name and isinstance(typology_name, str):
                    try:
                        bag_price = float(ws.cell(row=row, column=5).value)
                        self.typologies[typology_name.strip()] = Typology(name=typology_name.strip(), bag_price=bag_price)
                    except (TypeError, ValueError):
                        pass
                # Rates (post‑traitement, finition, TVA) live in column 7 and 8 on rows with type pieces
                rate_label = ws.cell(row=row, column=7).value
                rate_value = ws.cell(row=row, column=8).value
                if rate_label and isinstance(rate_label, str):
                    lower = rate_label.strip().lower()
                    try:
                        val = float(rate_value)
                    except (TypeError, ValueError):
                        val = None
                    if val is not None:
                        if 'post' in lower:
                            self.post_rate = val
                        elif 'finition' in lower:
                            self.finish_rate = val
                        elif 'tva' in lower:
                            self.tva_rate = val
                row += 1
            # After finishing type pieces, parse shipping costs explicitly
            for r in range(1, ws.max_row + 1):
                label = ws.cell(row=r, column=7).value
                value = ws.cell(row=r, column=8).value
                if not label:
                    continue
                if isinstance(label, str):
                    l = label.strip().lower()
                    try:
                        v = float(value)
                    except (TypeError, ValueError):
                        continue
                    if 'frais retrait' in l:
                        self.shipping_retrait = v
                    elif 'frais livraison' in l:
                        self.shipping_delivery = v
            # Find row containing 'MARK-UP (selon ml plateau)' and packaging
            for r in range(1, ws.max_row + 1):
                cell = ws.cell(row=r, column=1).value
                if isinstance(cell, str) and 'MARK-UP' in cell:
                    # The header row for mark-ups is at r+1
                    mrow = r + 1
                    # iterate rows until an empty cell
                    cur = mrow + 1
                    while True:
                        ml_min = ws.cell(row=cur, column=1).value
                        factor = ws.cell(row=cur, column=2).value
                        if ml_min is None or factor is None:
                            break
                        self.markup_table.append((float(ml_min), float(factor)))
                        cur += 1
                    # Packaging columns start at column 4 and 5 on the same rows
                    cur = mrow + 1
                    while True:
                        ml_total_min = ws.cell(row=cur, column=4).value
                        price_carton = ws.cell(row=cur, column=5).value
                        if ml_total_min is None or price_carton is None:
                            break
                        self.packaging_table.append((float(ml_total_min), float(price_carton)))
                        cur += 1
                    # Shipping delivery cost is in column 8 of the header row +1 (same row as
                    # packaging header). We loaded shipping_retrait earlier; next row holds shipping delivery.
                    try:
                        self.shipping_delivery = float(ws.cell(row=r, column=8).value or self.shipping_delivery)
                    except Exception:
                        pass
                    break
            # Sort the tables for binary search later
            self.markup_table.sort(key=lambda x: x[0])
            self.packaging_table.sort(key=lambda x: x[0])
            self._loaded = True

    def _get_markup_factor(self, volume_ml: float) -> float:
        """Return the mark‑up factor for a given volume (ml)."""
        factor = 1.0
        for ml_min, f in self.markup_table:
            if volume_ml >= ml_min:
                factor = f
            else:
                break
        return factor

    def _get_packaging_cost(self, total_volume_ml: float) -> float:
        """Return the packaging cost based on total volume (ml)."""
        price = 0.0
        for ml_min, p in self.packaging_table:
            if total_volume_ml >= ml_min:
                price = p
            else:
                break
        return price

    def compute_quote(
        self,
        volume_ml: float,
        material_name: str,
        type_piece_name: str,
        typology_name: str,
        quantity: int = 1,
        shipping_method: str = 'retrait',
        add_painting: bool = False,
        largest_dimension_mm: Optional[float] = None,
    ) -> Dict[str, float]:
        """Compute a detailed quote.

        Args:
            volume_ml: Volume of a single part in millilitres.
            material_name: Name of the material to use.
            type_piece_name: Category of the part (affects factor).
            typology_name: Typology for bag cost (affects packaging).
            quantity: Number of identical parts.
            shipping_method: 'retrait' for pickup or 'livraison' for delivery.
            add_painting: Whether painting is requested (adds finish cost and a
                placeholder painting fee). Painting fee is not defined in the
                Excel and must be implemented manually if needed.

        Returns:
            A dictionary containing the breakdown: material_cost, machine_cost,
            base_cost, post_cost, finish_cost, total_cost_before_markup,
            markup_factor, price_ht_plate, packaging_cost, bag_cost,
            shipping_cost, total_ht, vat, total_ttc.
        """
        if quantity < 1:
            quantity = 1
        # Retrieve material price
        if material_name not in self.materials:
            raise ValueError(f"Unknown material: {material_name}")
        mat = self.materials[material_name]
        # Retrieve type piece factor
        type_factor = self.type_pieces.get(type_piece_name, TypePiece(type_piece_name, 1.0)).factor
        # Retrieve typology bag price
        bag_price = self.typologies.get(typology_name, Typology(typology_name, 0.0)).bag_price
        # Effective volume with quantity and type factor (type factor scales time and material)
        eff_volume_ml = volume_ml * quantity
        # Add supports based on material
        support_percent = self.material_support_percent.get(material_name, 0.0)
        volume_with_supports_ml = eff_volume_ml * (1.0 + support_percent / 100.0)
        # Material cost
        material_cost = volume_with_supports_ml * mat.price_per_ml_with_loss * type_factor
        # Machine cost
        # Estimate print time in hours: using either machine_time_per_ml or bounding box + speed/time_factor
        machine_time_hours: float
        if largest_dimension_mm is not None:
            # Compute printing speed in mm/s for the material
            speed_tenth = self.material_print_speed.get(material_name)
            if speed_tenth is None:
                speed_mm_s = 1.0  # default to 1 mm/s
            else:
                speed_mm_s = speed_tenth / 10.0
            # Estimate time in seconds
            time_seconds = (largest_dimension_mm * self.time_factor) / speed_mm_s
            machine_time_hours = time_seconds / 3600.0
        else:
            # Fallback to volume based approximation
            machine_time_hours = volume_with_supports_ml * self.machine_time_per_ml * type_factor
        machine_cost = machine_time_hours * self.machine_hour_rate
        # Base cost
        base_cost = material_cost + machine_cost
        # Post‑treatment and finishing
        post_cost = base_cost * self.post_rate
        finish_cost = base_cost * self.finish_rate
        # Additional painting cost (set to zero as placeholder)
        painting_cost = 0.0
        if add_painting:
            # TODO: implement painting cost logic; for now, a flat 0
            painting_cost = 0.0
        total_cost_before_markup = base_cost + post_cost + finish_cost + painting_cost
        # Mark‑up factor based on per‑plate volume (assuming one plate per order)
        markup_factor = self._get_markup_factor(eff_volume_ml)
        price_ht_plate = total_cost_before_markup * markup_factor
        # Packaging and bag costs
        packaging_cost = self._get_packaging_cost(eff_volume_ml)
        bag_cost = bag_price * quantity
        # Shipping cost
        shipping_cost = self.shipping_retrait if shipping_method == 'retrait' else self.shipping_delivery
        # Total HT
        total_ht = price_ht_plate + packaging_cost + bag_cost + shipping_cost
        # VAT
        vat = total_ht * self.tva_rate
        total_ttc = total_ht + vat
        # Print time in minutes
        print_time_minutes = machine_time_hours * 60.0
        return {
            'material_cost': material_cost,
            'machine_cost': machine_cost,
            'base_cost': base_cost,
            'post_cost': post_cost,
            'finish_cost': finish_cost,
            'painting_cost': painting_cost,
            'total_cost_before_markup': total_cost_before_markup,
            'markup_factor': markup_factor,
            'price_ht_plate': price_ht_plate,
            'packaging_cost': packaging_cost,
            'bag_cost': bag_cost,
            'shipping_cost': shipping_cost,
            'total_ht': total_ht,
            'vat': vat,
            'total_ttc': total_ttc,
            'volume_with_supports_ml': volume_with_supports_ml,
            'print_time_minutes': print_time_minutes,
        }
